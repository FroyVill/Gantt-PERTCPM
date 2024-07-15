import tkinter as tk
from tkinter import ttk, messagebox
from matplotlib import patches
from matplotlib.widgets import Button
import tkinter 
import os
import openpyxl
import matplotlib.pyplot as plt
from sklearn import tree
from tkcalendar import DateEntry
import customtkinter
import matplotlib.pyplot as plt
import pandas as pd
import networkx as nx


def PERT_CPM_button_clicked():
    show_PERT_CPM_window()

#Path
filepath = os.path.join("C:", os.sep, "#", "#", "#", "#", "#", "Interfaz", "ExcelPERT.xlsx")


def calcular_t_estimado(t_optimista, t_probable, t_pesimista):
    return (t_optimista + (4 * t_probable) + t_pesimista) / 6

def enter_data_PERT_CPM():
    global nodo_in, act_in, predescesor_in, tiempo_opt_in, tiempo_m_prob_in, tiempo_pesimista_in, tiempo_est_in

    # Obtención de datos
    clave = nodo_in.get()
    actividad = act_in.get()
    predescesor = predescesor_in.get() if predescesor_in.get() else "inicio"  

    if clave and actividad:
        try:
            if "," in predescesor:
                predescesores = [p.strip() for p in predescesor.split(",")]
            else:
                #Si es solo un predecesor, lo dejamos como una lista de un solo elemento
                predescesores = [predescesor.strip()]
            t_pesimista = float(tiempo_pesimista_in.get()) if tiempo_pesimista_in.get() else 0 
            t_probable = float(tiempo_m_prob_in.get()) if tiempo_m_prob_in.get() else 0
            t_optimista = float(tiempo_opt_in.get()) if tiempo_opt_in.get() else 0
        except ValueError:
            messagebox.showerror(title="Error", message="Los tiempos tienen que ser solamente números")
            return

        # Calcular el tiempo estimado
        t_estimado = calcular_t_estimado(t_optimista, t_probable, t_pesimista)


        #Si no existe el excel se crea uno
        if not os.path.exists(filepath):
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            #Titulo de cada columna
            heading = ["Clave", "Actividad", "Predescesora", "Tiempo optimista", "Tiempo más probable", "Tiempo pesimista", "Tiempo estimado"]
            sheet.append(heading)
        else:
            #Si existe el excel cargarlo
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active

        #Agregar los datos
        for predescesor in predescesores:
            sheet.append([clave, actividad, predescesor, t_optimista, t_probable, t_pesimista, t_estimado])

        #Guardar
        workbook.save(filepath)

        messagebox.showinfo(title="Éxito", message="Datos guardados correctamente")
    else:
        messagebox.showwarning(title="Error", message="Falta ingresar datos")


def delete_data_PERT_CPM():
    if os.path.exists(filepath):
        try:
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            sheet.delete_rows(2, sheet.max_row)
            workbook.save(filepath)
            messagebox.showinfo(title="Exito", message="Datos borrados correctamente.")
        except Exception as e:
            messagebox.showerror(title="Error", message=f"No se pudo abrir el archivo: {e}")
    else:
        messagebox.showwarning(title="Error", message="No se encontró el archivo.")

def edit_data_button_clicked_PERT_CPM():
    edit_data_PERT_CPM()

def edit_data_PERT_CPM():
    def edit_act():
        start_a = selected_id.get()
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value == start_a:
                    row_number = cell.row
                    end = act.get()  
                    ws[f"B{row_number}"].value = end  #Seleccionar la columna "B"
        wb.save(filepath) 
        status_label.config(text="Actividad actualizada exitosamente!!")
    
    def edit_nodo_p():
        start_n = selected_id.get()
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value == start_n:
                    row_number = cell.row
                    end = nodo_pred.get()  
                    ws[f"C{row_number}"].value = end  #Seleccionar la columna "C"
        wb.save(filepath) 
        status_label.config(text="Nodo predescesor actualizado exitosamente!!")
    
    def edit_t_opt():
        start_to = selected_id.get()
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value == start_to:
                    row_number = cell.row
                    end = tiempo_opt.get()  
                    ws[f"D{row_number}"].value = end  #Seleccionar la columna "D"
        wb.save(filepath) 
        status_label.config(text="Tiempos optimista actualizado exitosamente!!")

    def edit_t_m_prob():
        start_tmp = selected_id.get()
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value == start_tmp:
                    row_number = cell.row
                    end = tiempo_m_p.get()  
                    ws[f"E{row_number}"].value = end  #Seleccionar la columna "E"
        wb.save(filepath) 
        status_label.config(text="El tiempo más probable se ha actualizado exitosamente!!")

    def edit_t_pes():
        start_tp = selected_id.get()
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value == start_tp:
                    row_number = cell.row
                    end = tiempo_pes.get()  
                    ws[f"F{row_number}"].value = end  #Seleccionar la columna "F"
        wb.save(filepath) 
        status_label.config(text="Tiempos pesimista se ha actualizado exitosamente!!")

    wb = openpyxl.load_workbook(filepath)
    ws = wb["Sheet"]

    #Crear frame principal
    root = tk.Tk()
    root.title("Editar datos")
    root.resizable(False,False)

    #Añadir frame principal a custom tkinter
    frame2 = customtkinter.CTkFrame(root,fg_color="#E5E5E5")
    frame2.pack()
    root.eval('tk::PlaceWindow . center')

    #Crear frame para seleccionar
    select_frame = customtkinter.CTkFrame(frame2,fg_color="#D9D9D9")
    select_frame.grid(row=0, column=0, padx=20, pady=10)
    
    #--------------Nodo-------------
    #Recopilar los valores de las celdas en la primera columna 
    tasks = [cell.value for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1) for cell in row]

    selected_id = tk.StringVar(value=tasks[0])

    #Etiqueta seleccionar
    select_label = customtkinter.CTkLabel(select_frame, text="Seleccionar la actividad",font=('Helvetica', 20),text_color="black")
    select_label.grid(row=0, column=0,padx=10)

    #Spinbox para seleccionar una etiqueta ya hecha
    select_spinbox=ttk.Spinbox(select_frame, value=tasks,width=10, textvariable=selected_id)
    select_spinbox.grid(row=1, column=0, padx=5, pady=5)

    #-------------------Act y nodo prescedente--------
    #Etiquetas de las actividades
    act_label = customtkinter.CTkLabel(select_frame, text="Actividad",font=('Helvetica', 20),text_color="black")
    act_label.grid(row=0, column=1, padx= 20, pady=10)
    nodo_pred_label = customtkinter.CTkLabel(select_frame, text="Nodo predescesor",font=('Helvetica', 20),text_color="black")
    nodo_pred_label.grid(row=0, column=2, padx=20, pady=10)

    #Entrada de los datos
    act = customtkinter.CTkEntry(select_frame,fg_color='white',border_color="#9D9EA0",text_color="black")
    act.grid(row=1, column=1, padx=20, pady=10)
    nodo_pred = customtkinter.CTkEntry(select_frame,fg_color='white',border_color="#9D9EA0",text_color="black")
    nodo_pred.grid(row=1, column=2, padx=20, pady=10)

    #Botones para actualizar los datos
    edit_act_btn = customtkinter.CTkButton(select_frame, text="Actualizar",font=('Helvetica', 20),text_color="white",fg_color='#284EEE', command=edit_act)
    edit_act_btn.grid(row=2, column=1, padx=20, pady=(5,10))
    edit_nodo_p_btn = customtkinter.CTkButton(select_frame, text="Actualizar",font=('Helvetica', 20),text_color="white",fg_color='#284EEE', command=edit_nodo_p)
    edit_nodo_p_btn.grid(row=2, column=2, padx=20, pady=(5,10))
    #-------------------Tiempos-----------------
    #Frame de los tiempos
    new_tiempos_frame = customtkinter.CTkFrame(frame2,fg_color="#D9D9D9")
    new_tiempos_frame.grid(row=2, column=0, sticky="news", padx=50, pady=10)
    
    #Etiqueta de los nuevos tiempos
    tiempo_opt_label = customtkinter.CTkLabel(new_tiempos_frame, text="Tiempo optimista", font=('Helvetica', 20),text_color="black")
    tiempo_opt_label.grid(row=4, column=0, padx=20, pady=10)
    tiempo_m_prob_label = customtkinter.CTkLabel(new_tiempos_frame, text="Tiempo más probable", font=('Helvetica', 20),text_color="black")
    tiempo_m_prob_label.grid(row=4, column=1, padx=20, pady=10)
    tiempo_pesimista_label = customtkinter.CTkLabel(new_tiempos_frame, text="Tiempo pesimista",font=('Helvetica', 20),text_color="black")
    tiempo_pesimista_label.grid(row=4, column=2, padx=20, pady=10)

    #Escoger los tiempos
    tiempo_opt = customtkinter.CTkEntry(new_tiempos_frame,fg_color='white',border_color="#9D9EA0",text_color="black")
    tiempo_opt.grid(row=5, column=0, padx=20, pady=10)
    tiempo_m_p = customtkinter.CTkEntry(new_tiempos_frame,fg_color='white',border_color="#9D9EA0",text_color="black")
    tiempo_m_p.grid(row=5, column=1, padx=20, pady=10)
    tiempo_pes = customtkinter.CTkEntry(new_tiempos_frame,fg_color='white',border_color="#9D9EA0",text_color="black")
    tiempo_pes.grid(row=5, column=2, padx=20, pady=10)

    #Botones para guardar los cambios
    edit_t_opt_btn = customtkinter.CTkButton(new_tiempos_frame, text="Actualizar",font=('Helvetica', 20),text_color="white",fg_color='#284EEE', command=edit_t_opt)
    edit_t_opt_btn.grid(row=6, column=0, padx=20, pady=(5,10))
    edit_t_m_prob_btn = customtkinter.CTkButton(new_tiempos_frame, text="Actualizar",font=('Helvetica', 20),text_color="white",fg_color='#284EEE', command=edit_t_m_prob)
    edit_t_m_prob_btn.grid(row=6, column=1, padx=20, pady=(5,10))
    edit_t_pes_btn = customtkinter.CTkButton(new_tiempos_frame, text="Actualizar",font=('Helvetica', 20),text_color="white",fg_color='#284EEE', command=edit_t_pes)
    edit_t_pes_btn.grid(row=6, column=2, padx=20, pady=(5,10))

    #-------------------------------------------------
    # Frame de los botones auxiliares------------------------------------
    aux_frame2 = customtkinter.CTkFrame(frame2,fg_color="#D9D9D9")
    aux_frame2.grid(row=3, column=0, padx=20, pady=10)

    volver_btn = customtkinter.CTkButton(aux_frame2, text="Volver",font=('Helvetica', 20,'bold'),height=40,fg_color='#3A7EBF', command=lambda: [root.destroy(), show_PERT_CPM_window()])
    volver_btn.grid(row=0, column=0, padx=20, pady=(10, 10))

    #Mostrar mensaje de exito--------------------
    status_label = ttk.Label(root, text="")
    status_label.pack(pady=10)

    root.mainloop()

def pert_cpm():

    if os.path.exists(filepath):
        try:
            df = pd.read_excel(filepath)
        except Exception as e:
            print(f"No se pudo abrir el archivo: {e}")
            return
    else:
        print("No se encontró el archivo.")
        return

    #Crear el gráfico PERT-CPM
    G = nx.DiGraph()

    # Añadir el nodo "inicio"
    G.add_node("inicio", label="Inicio")

    for index, row in df.iterrows():
        G.add_node(row['Clave'], label=row['Actividad'])  

        if isinstance(row['Predescesora'], str):  #Verificar si es una cadena
            predescesores = [p.strip() for p in row['Predescesora'].split(",")]
            #Conectamos el nodo actual con cada uno de sus predecesores
            for predecesor in predescesores:
                if predecesor:  #Verificar que el predecesor no esté vacío
                    G.add_edge(predecesor, row['Clave'])

    #Usar kamada_kawai_layout para calcular las posiciones de los nodos
    pos = nx.kamada_kawai_layout(G)

    labels = nx.get_node_attributes(G, 'label')
    plt.close('all')
    plt.figure(figsize=(15, 10))  #Aumentar el tamaño de la figura

    #Dibujar los nodos con las etiquetas adentro
    nx.draw_networkx_nodes(G, pos, node_color='lightblue', node_size=2500)

    #Dibujar el nombre del nodo adentro del círculo
    for node, (x, y) in pos.items():
        label = labels[node]
        #if node != "inicio":  #Evitar repetir el nombre de la actividad para el nodo "inicio"
         #   label = f"{node}: {label}"  
        plt.text(x, y, label, fontsize=10, ha='center', va='center')

    #Dibujar las aristas con flechas y etiquetas
    for edge in G.edges():
        nx.draw_networkx_edges(G, pos, edgelist=[edge], width=1.0, alpha=0.5, arrows=True)

        #Calculamos el punto medio de la arista para poner la etiqueta de la flecha
        x0, y0 = pos[edge[0]]
        x1, y1 = pos[edge[1]]
        xm, ym = (x0 + x1) / 2, (y0 + y1) / 2

        #Dibujamos la etiqueta de la flecha
        plt.annotate("", xy=(x1, y1), xytext=(x0, y0), arrowprops=dict(arrowstyle="->"))
        
        #Dibujamos el número de nodo en el medio de la flecha, con negrita y más arriba de la flecha
        plt.text(xm, ym + 0.1, edge[1], fontsize=10, ha='center', va='center', fontweight='bold')

    #Identificar la ruta crítica
    try:
        critical_path = nx.dag_longest_path(G)
        critical_edges = [(critical_path[i], critical_path[i+1]) for i in range(len(critical_path)-1)]

        #Dibujar las aristas de la ruta crítica
        nx.draw_networkx_edges(G, pos, edgelist=critical_edges, width=2.0, alpha=0.7, edge_color='red', arrows=False)

        #Dibujar la flecha de la ruta crítica en rojo
        for edge in critical_edges:
            plt.annotate("", xy=pos[edge[1]], xytext=pos[edge[0]], arrowprops=dict(arrowstyle="->", color='red'))
        
    except nx.NetworkXError:
        pass  #En caso de que no haya camino crítico

    plt.title('Diagrama PERT-CPM')

    def on_button_clicked(event):
        #Calcular los tiempos estimados de cada nodo
        df['Tiempo estimado'] = df.apply(lambda row: calcular_t_estimado(row['Tiempo optimista'], row['Tiempo más probable'], row['Tiempo pesimista']), axis=1)
    
        #Inicializar FIP y FTP de cada nodo con valores vacíos
        df['FIP'] = ""
        df['FTP'] = ""
        df['Predescesor'] = df['Predescesora']  # Crear una nueva columna 

        #Iterar sobre cada nodo, empezando desde el segundo
        for index, row in df.iterrows():
            if row['Clave'] != 'inicio':  # Evitar el nodo "inicio"
                predescesores = row['Predescesora'].split(",") if isinstance(row['Predescesora'], str) else []

                #Calcular el valor máximo de FIP y FTP entre los predecesores
                max_FIP = max(df[df['Clave'].isin(predescesores)]['FIP'], default=0)
                max_FTP = max(df[df['Clave'].isin(predescesores)]['FTP'], default=0)

                #Actualizar los valores de FIP y FTP del nodo actual
                df.at[index, 'FIP'] = max(max_FIP, max_FTP)
                df.at[index, 'FTP'] = max_FTP + row['Tiempo estimado']

        #Se establece FIP=0 solo para el primer nodo
        df.loc[df['Clave'] == 'inicio', 'FIP'] = 0

        # Calular FIL y FTL
        # Calcular FTL para el último nodo
        last_node = df['Clave'].iloc[-1]
        df.loc[df['Clave'] == last_node, 'FTL'] = max(df.loc[df['Clave'] == last_node, ['FIP', 'FTP']].values.flatten())

        #Calcular FIL para todos los nodos
        for index, row in df.iterrows():
            max_FIP_FTP = max(row['FIP'], row['FTP'])
            df.at[index, 'FIL'] = max_FIP_FTP - row['Tiempo estimado']

        #Calcular FTL y FIL para los nodos restantes
        for index, row in df.iloc[:-1].iterrows():
            #Tomar el número más chico entre FIL y FTL de los nodos sucesores
            min_FIL_FTL_sucesores = min(df.loc[df['Predescesor'].str.contains(row['Clave']), ['FIL', 'FTL']].values.flatten(), default=0)
            #Actualizar FTL del nodo actual con el valor mínimo encontrado
            df.at[index, 'FTL'] = min_FIL_FTL_sucesores
        

        for index, row in df.iterrows():
            df.at[index, 'FIL'] = row['FTL'] - row['Tiempo estimado']
    
        #Calcular FTL y FIL para los nodos restantes
        for index, row in df.iloc[:-1].iterrows():
            #Tomar el número más chico entre FIL y FTL de los nodos sucesores
            min_FIL_FTL_sucesores = min(df.loc[df['Predescesor'].str.contains(row['Clave']), ['FIL', 'FTL']].values.flatten(), default=0)
            #Actualizar FTL del nodo actual con el valor mínimo encontrado
            df.at[index, 'FTL'] = min_FIL_FTL_sucesores

        for index, row in df.iterrows():
            df.at[index, 'FIL'] = row['FTL'] - row['Tiempo estimado']

        #Calcular FTL y FIL para los nodos restantes
        for index, row in df.iloc[:-1].iterrows():
            # Tomar el número más chico entre FIL y FTL de los nodos sucesores
            min_FIL_FTL_sucesores = min(df.loc[df['Predescesor'].str.contains(row['Clave']), ['FIL', 'FTL']].values.flatten(), default=0)
            #Actualizar FTL del nodo actual con el valor mínimo encontrado
            df.at[index, 'FTL'] = min_FIL_FTL_sucesores

        for index, row in df.iterrows():
            df.at[index, 'FIL'] = row['FTL'] - row['Tiempo estimado']

        #Calcular la diferencia FTP - FIL para cada nodo
        df['Diferencia'] = df['FIL'] - df['FIP']

        #Mostrar los resultados en una ventana emergente
        window = tk.Tk()
        window.title("Tablas de Holgura")

        frame = tk.Frame(window)
        frame.pack(padx=20, pady=10)

        tree = ttk.Treeview(frame, columns=("Nodo",  "Actividad", "Predecesor", "Tiempo Estimado", "FIP", "FTP", "FIL", "FTL", "Diferencia"), show="headings")

        #Definir el ancho y centrar texto para todas las columnas
        column_width = 100
        alignment = 'center'
        for column in ("Nodo",  "Actividad", "Predecesor", "Tiempo Estimado", "FIP", "FTP", "FIL", "FTL", "Diferencia"):
            tree.heading(column, text=column, anchor=alignment)
            tree.column(column, width=column_width, anchor=alignment)

        for index, row in df.iterrows():
            tree.insert("", "end", values=(row['Clave'], row['Actividad'], row['Predescesora'], f"{row['Tiempo estimado']:.0f}", f"{row['FIP']:.0f}", f"{row['FTP']:.0f}", f"{row['FIL']:.0f}", f"{row['FTL']:.0f}", f"{row['Diferencia']:.0f}"))

        tree.pack()

    #Añadir el fondo azul detrás del botón
    ax_button = plt.axes([0.73, 0.12, 0.2, 0.075])  # Coordenadas [left, bottom, width, height]
    rect = patches.Rectangle((0.25, 0.25), 0.5, 0.5, transform=ax_button.transAxes, color='blue', zorder=-1)
    ax_button.add_patch(rect)

    #Crear el botón en grafica
    button = Button(ax_button, 'Tabla de holgura')
    button.label.set_color('white')
    button.on_clicked(on_button_clicked)

    plt.axis('off')
    plt.show()






def show_PERT_CPM_window():
    global nodo_in, act_in, predescesor_in, tiempo_opt_in, tiempo_m_prob_in, tiempo_pesimista_in, tiempo_est_in

    #Frame principal
    window = tk.Tk()
    window.title("Pert/CPM")
    window.resizable(False,False)

    #Añadir frame principal a customtkinker
    frame = customtkinter.CTkFrame(window,fg_color="#E5E5E5")
    frame.pack()
    

    #Posicionar ventana en medio de la pantalla (Funciona a medias)
    window.eval('tk::PlaceWindow . center')
    #customtkinter.set_appearance_mode("black")
    #customtkinter.set_default_color_theme("dark-blue")

    # Hacer pantalla completa
    #window.attributes('-fullscreen', True)

    #Frame de actividades
    TakAct_frame = customtkinter.CTkFrame(frame,fg_color="#D9D9D9")
    TakAct_frame.grid(row=0, column=0, padx=20, pady=10)

    #contenido_frame = customtkinter.CTkFrame(frame)
    #contenido_frame.grid(row=2, column=0)
    

    #Etiqueta de Nodo
    nodo_label = customtkinter.CTkLabel(TakAct_frame, text="Nodo",font=('Helvetica', 20),text_color="black")
    nodo_label.grid(row=0, column=0)

    #Etiqueta de actividad
    act_label = customtkinter.CTkLabel(TakAct_frame, text="Actividad",font=('Helvetica', 20),text_color="black")
    act_label.grid(row=0, column=1)

    #Etiqueta del nodo predecesor
    pred_label = customtkinter.CTkLabel(TakAct_frame, text="Nodo predescesor",font=('Helvetica', 20),text_color="black")
    pred_label.grid(row=0, column=2)

    #Entry para escribir el nodo, escribir la actividad y el nodo predescesor
    nodo_in = customtkinter.CTkEntry(TakAct_frame,fg_color='white',border_color="#9D9EA0",text_color="black")
    act_in = customtkinter.CTkEntry(TakAct_frame,fg_color='white',border_color="#9D9EA0",text_color="black")
    predescesor_in =customtkinter.CTkEntry(TakAct_frame,fg_color='white',border_color="#9D9EA0",text_color="black")
    nodo_in.grid(row=1, column=0)
    act_in.grid(row=1, column=1)
    predescesor_in.grid(row=1,column=2)

    #Mejora la aparencia y la disposición general de los widgets
    for widget in TakAct_frame.winfo_children():
        widget.grid_configure(padx=10, pady=5)

    #Frame de los tiempos
    tiempos_frame = customtkinter.CTkFrame(frame,fg_color="#D9D9D9")
    tiempos_frame.grid(row=1, column=0, sticky="news", padx=20, pady=10)

    #Etiqueta de los tiempos
    t_optimista_label = customtkinter.CTkLabel(tiempos_frame, text="Tiempo optimista",font=('Helvetica', 20),text_color="black")
    t_m_prob_label = customtkinter.CTkLabel(tiempos_frame, text="Tiempo más probable",font=('Helvetica', 20),text_color="black")
    t_pesimista_label = customtkinter.CTkLabel(tiempos_frame, text="Tiempo pesimista",font=('Helvetica', 20),text_color="black")
    t_optimista_label.grid(row=0,column=0)
    t_m_prob_label.grid(row=0, column=1)
    t_pesimista_label.grid(row=0, column=2)

    #Entry de los tiempos
    tiempo_opt_in = customtkinter.CTkEntry(tiempos_frame,fg_color='white',border_color="#9D9EA0",text_color="black")
    tiempo_m_prob_in = customtkinter.CTkEntry(tiempos_frame,fg_color='white',border_color="#9D9EA0",text_color="black")
    tiempo_pesimista_in = customtkinter.CTkEntry(tiempos_frame,fg_color='white',border_color="#9D9EA0",text_color="black")
    tiempo_opt_in.grid(row=1, column=0)
    tiempo_m_prob_in.grid(row=1, column=1)
    tiempo_pesimista_in.grid(row=1, column=2)

    for widget in tiempos_frame.winfo_children():
        widget.grid_configure(padx=30, pady=15)

    #Frame de los botones
    btn_frame = customtkinter.CTkFrame(frame,fg_color="#D9D9D9")
    btn_frame.grid(row=2, column=0, padx=20,  pady=(20, 20))

    #Boton para guardar
    btn_save = customtkinter.CTkButton(btn_frame,fg_color='#284EEE', text="Guardar datos",height=40, font=('Helvetica', 20),command=enter_data_PERT_CPM)
    btn_save.grid(row=0, column=0, sticky="news", padx=20, pady=(10, 10))
    
    #Boton para editar
    btn_edit = customtkinter.CTkButton(btn_frame,fg_color='#284EEE', text="Editar datos",height=40, font=('Helvetica', 20), command=lambda: [window.destroy(), edit_data_button_clicked_PERT_CPM()])
    btn_edit.grid(row=0, column=1, sticky="news", padx=20,  pady=(10, 10))
    
    #Boton para borrar
    btn_del = customtkinter.CTkButton(btn_frame,fg_color='#284EEE', text="Borrar datos",height=40, font=('Helvetica', 20), command=delete_data_PERT_CPM)
    btn_del.grid(row=0, column=2, sticky="news", padx=20,  pady=(10, 10))


    aux2_frame = customtkinter.CTkFrame(frame,fg_color="#D9D9D9")
    aux2_frame.grid(row=4, column=0, padx=20,  pady=(10, 10))

    btn_pert = customtkinter.CTkButton(aux2_frame,fg_color='#284EEE', text="Graficar", command=pert_cpm,height=40, font=('Helvetica', 20))
    btn_pert.grid(row=0, column=0, sticky="news", padx=20, pady=10)
    
    from gantt import Gantt_button_clicked
    volver_btn = customtkinter.CTkButton(aux2_frame,fg_color='#3A7EBF', text="GANTT", height=40, font=('Helvetica', 20),command=lambda: [window.destroy(), Gantt_button_clicked()])
    volver_btn.grid(row=0, column=2, sticky="news", padx=20,  pady=(10, 10))

    # Botón para salir
    buttonSalir = customtkinter.CTkButton(aux2_frame, text="Salir", height=40, font=('Helvetica', 20, 'bold'), fg_color='red', hover_color='#B80507', command=window.destroy)
    buttonSalir.grid(row=1, column=1, padx=20, pady=(10, 10))

    # Centrar el contenido_frame en la ventana
    #contenido_frame.place(relx=0.5, rely=0.5, anchor='center')

    window.mainloop()
