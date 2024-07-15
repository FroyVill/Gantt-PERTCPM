import tkinter as tk
from tkinter import ttk, messagebox
import os
import openpyxl
from tkcalendar import DateEntry
import customtkinter
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.patches import Patch
import pandas as pd
import networkx as nx
import random
from cpm import PERT_CPM_button_clicked
from tkinter import *
from datetime import datetime


def Gantt_button_clicked():
    show_Gantt_window()

#path
filepath = os.path.join("C:", os.sep, "#", "#", "#", "#", "#","Interfaz", "ExcelGantt.xlsx")

#Guardar datos
def enter_data_GANTT():
    global task_entry, department_entry, start_entry, end_entry, completion_scale
    
    #Obtener datos de tareas y departamentos
    task = task_entry.get()
    department = department_entry.get()
    
    if task and department:
        #Obtener datos de fechas y porcentajes
        start_status = start_entry.get_date().strftime('%d/%m/%Y')
        end_status = end_entry.get_date().strftime('%d/%m/%Y')
        completion = float(completion_scale.get())
        
        #Crear excel si es que no esxiste uno
        if not os.path.exists(filepath):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            #Poner encabezados 
            heading = ["Task", "Department", "Start", "End", "Completion"]
            sheet.append(heading)
        else:
            #Si existe cargarlo
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
        
        #Obtener la última fila con datos
        last_row = sheet.max_row

        #Asegurarse de que no hay filas vacías
        while sheet.cell(row=last_row, column=1).value is None and last_row > 1:
            last_row -= 1

        #Agregar los datos en la siguiente fila disponible
        sheet.append([task, department, start_status, end_status, completion])

        #Guardar
        workbook.save(filepath)
        
        #Mensaje de éxito
        tk.messagebox.showinfo(title="Exito", message="Datos guardados correctamente")
            
    else:
        tk.messagebox.showwarning(title="Error", message="Falta ingresar datos")
def delete_data_GANTT():      
    if os.path.exists(filepath):
        try:
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active
            #Eliminar todas las filas del excel desde la fila 2 hasta la última fila
            sheet.delete_rows(2, sheet.max_row)
            workbook.save(filepath)#Guardar cambios hechos
            tk.messagebox.showinfo(title="Exito", message="Datos borrados correctamente.")
        except Exception as e:
            tk.messagebox.showerror(title="Error", message=f"No se pudo abrir el archivo: {e}")
    else:
        tk.messagebox.showwarning(title="Error", message="No se encontró el archivo.")

def edit_data_button_clicked_GANTT():
    edit_data_GANTT()

def edit_data_GANTT():
    def update_completion():
        task = selected_task.get()  
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value == task:
                    row_number = cell.row
                    ws[f"E{row_number}"].value = completion_value.get()  #Seleccionar la fila "E"
        wb.save(filepath)  
        status_label.config(text="Porcentaje actualizado exitosamente!!")

    def edit_fecha():
        start = selected_task.get()  
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value == start:
                    row_number = cell.row
                    end = fecha_entry.get()  
                    ws[f"D{row_number}"].value = end  #Seleccionar la fila "D"
        wb.save(filepath) 
        status_label.config(text="Fecha actualizada exitosamente!!")

    #Cargar excel
    wb = openpyxl.load_workbook(filepath)
    ws = wb["Sheet"]

    #Crear frame principal
    root = tk.Tk()
    root.title("Editar datos")
    root.resizable(False,False)

    #Añadir frame principal a customtkinter
    frame2 = customtkinter.CTkFrame(root,fg_color="#E5E5E5")
    frame2.pack()
    root.eval('tk::PlaceWindow . center')


    #Crear frame para seleccionar
    select_frame = customtkinter.CTkFrame(frame2,fg_color="#D9D9D9")
    select_frame.grid(row=0, column=0, padx=20, pady=10) 

    #Recopilar los valores de las celdas en la primera columna
    tasks = [cell.value for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1) for cell in row]

    selected_task = tk.StringVar(value=tasks[0])

    #Etiqueta de seleccionar tarea
    select_label = customtkinter.CTkLabel(select_frame, text="Seleccionar Tarea",font=('Helvetica', 20),text_color="black")
    select_label.grid(row=0, column=0,padx=(5, 20), pady=(5, 20))

    #Spinbox para seleccionar una tarea ya hecha
    select_spinbox = ttk.Spinbox(select_frame, values=tasks, textvariable=selected_task)
    select_spinbox.grid(row=1, column=0, padx=5, pady=5)

    # ----------------Fecha---------------
    #Frame de fecha
    Newdatos_frame = customtkinter.CTkFrame(frame2,fg_color="#D9D9D9")
    Newdatos_frame.grid(row=1, column=0, sticky="news", padx=50, pady=20)

    #Etiqueta de Nueva fecha final
    fecha_label = customtkinter.CTkLabel(Newdatos_frame, text="Nueva fecha final:",font=('Helvetica', 20),text_color="black")
    fecha_label.grid(row=0, column=0, padx=(5, 20), pady=(5, 0))

    #Escoger fecha
    fecha_entry = DateEntry(Newdatos_frame, date_pattern="dd/mm/yyyy" ,width=12, foreground='white', borderwidth=2, background="#1F538D",)
    fecha_entry.grid(row=1, column=0,padx=(5, 20), pady=(5, 0))

    #Boton para guardar nueva fecha
    edit_fecha_button = customtkinter.CTkButton(Newdatos_frame, text="Editar fecha", fg_color='#284EEE',command=edit_fecha, height=40,font=('Helvetica', 20),text_color="white")
    edit_fecha_button.grid(row=2, column=0, columnspan=2, padx=(5, 20), pady=(5, 20))

    # --------------Porcentaje de avance-----------

    #Etiqueta Nuevo porcentaje de avance
    completion_label = customtkinter.CTkLabel(Newdatos_frame, text="Nuevo % de avance:",font=('Helvetica', 20),text_color="black")
    completion_label.grid(row=0, column=2, padx=(5, 20), pady=(5, 0))
    completion_value = tk.DoubleVar()

    #Scale para seleccionar nuevo porcentaje
    completion_spinbox = tk.Scale(Newdatos_frame, from_=0.0, to=1.0, resolution=0.1, orient=tk.HORIZONTAL, variable=completion_value, background="#F9F9FB", fg="black", highlightbackground="#9D9EA0")
    completion_spinbox.grid(row=1, column=2,padx=(5, 20), pady=(5, 0))

    #Boton para Actualizar porcentaje de avance
    update_button = customtkinter.CTkButton(Newdatos_frame,fg_color='#284EEE',height=40, text="Actualizar %",font=('Helvetica', 20),text_color="white", command=update_completion)
    update_button.grid(row=2, column=2, columnspan=2, padx=(5, 20), pady=(5, 20))

    #Frame para volver-----------------
    volver_frame = customtkinter.CTkFrame(frame2,fg_color="#D9D9D9")
    volver_frame.grid(row=3, column=0, padx=20, pady=10) 

    volver_btn = customtkinter.CTkButton(volver_frame, text="Volver", fg_color='#3A7EBF',height=40, font=('Helvetica', 20, 'bold'), command=lambda: [root.destroy(), show_Gantt_window()])
    volver_btn.grid(row=0, column=0, padx=20, pady=(10, 10))

    #Mostrar mensaje de éxito
    status_label = ttk.Label(root, text="")
    status_label.pack(pady=10)

    root.mainloop()

def show_Gantt_window():
    global task_entry, department_entry, start_entry, end_entry, completion_scale

    #Frame principal
    window = tk.Tk()
    window.title("Diagrama Gantt")
    window.resizable(False,False)

    #Añadir frame principal a customtkinker
    frame = customtkinter.CTkFrame(window, fg_color="#E5E5E5")
    frame.pack()
    window.eval('tk::PlaceWindow . center')

    #Frame de departamentos
    TakDep_frame = customtkinter.CTkFrame(frame,fg_color="#D9D9D9")
    TakDep_frame.grid(row=0, column=0, padx=20, pady=10)

    #Etiqueta Actividad
    task_label = customtkinter.CTkLabel(TakDep_frame, text="Actividad",font=('Helvetica', 20),text_color="black")
    task_label.grid(row=0, column=0)

    #Etiqueta Responsable
    department_label = customtkinter.CTkLabel(TakDep_frame, text="Responsable",font=('Helvetica', 20),text_color="black")
    department_label.grid(row=0, column=1)

    #Entry para escribir la actividad y seleccionar el departamento
    task_entry = customtkinter.CTkEntry(TakDep_frame,fg_color='white',border_color="#9D9EA0",text_color="black")
    department_entry = customtkinter.CTkEntry(TakDep_frame,fg_color='white',border_color="#9D9EA0",text_color="black")
    task_entry.grid(row=1, column=0)
    department_entry.grid(row=1, column=1)

    for widget in TakDep_frame.winfo_children():
        widget.grid_configure(padx=10, pady=5)

    #Frame de fechas
    fechas_frame = customtkinter.CTkFrame(frame,fg_color="#D9D9D9")
    fechas_frame.grid(row=1, column=0, sticky="news", padx=20, pady=10)

    #Etiqueta de Inicia y y seleccionador de fecha
    start_label = customtkinter.CTkLabel(fechas_frame, text="Inicia",font=('Helvetica', 20),text_color="black")
    start_entry = DateEntry(fechas_frame, width=12, foreground='white', background="#1F538D")
    start_label.grid(row=0, column=0)
    start_entry.grid(row=1, column=0)

    #Etiqueta de Termina y seleccionador de fecha
    end_label = customtkinter.CTkLabel(fechas_frame, text="Termina",font=('Helvetica', 20),text_color="black")
    end_entry = DateEntry(fechas_frame, width=12, foreground='white', background="#1F538D")
    end_label.grid(row=0, column=1,pady=20)
    end_entry.grid(row=1, column=1,pady=20)

    #Etiqueta de Porcentaje y scale para seleccionar los numeros
    completion_label = customtkinter.CTkLabel(fechas_frame, text="Porcentaje de avance",text_color="black",font=('Helvetica', 20))
    completion_value = tk.DoubleVar()
    completion_scale = tk.Scale(fechas_frame,width=15, from_=0.0, to=1.0, resolution=0.1, orient=tk.HORIZONTAL, variable=completion_value, background="#F9F9FB", fg="black", highlightbackground="#9D9EA0")
    completion_label.grid(row=0, column=2)#pady=(5,10)
    completion_scale.grid(row=1, column=2)

    for widget in fechas_frame.winfo_children():
        widget.grid_configure(padx=30, pady=5)

    #Frame de los botones
    button_frame = customtkinter.CTkFrame(frame,fg_color="#D9D9D9")
    button_frame.grid(row=2, column=0, padx=20, pady=10)

    #Boton para guardar los datos
    button = customtkinter.CTkButton(button_frame, text="Guardar datos",font=('Helvetica', 20),height=40,fg_color='#284EEE', command=enter_data_GANTT) #Se llama funcion enter_data
    button.grid(row=0, column=0, sticky="news", padx=20, pady=10)

    #Boton para borrar los datos
    button2 = customtkinter.CTkButton(button_frame, text="Borrar datos",font=('Helvetica', 20),height=40,fg_color='#284EEE', command=delete_data_GANTT) #Se llama funcion delete data
    button2.grid(row=0, column=1, sticky="news", padx=20, pady=10)

    #Boton para editar los datos
    edit_button = customtkinter.CTkButton(button_frame, text="Editar", font=('Helvetica', 20),height=40,fg_color='#284EEE',command=lambda: [window.destroy(), edit_data_button_clicked_GANTT()])
    edit_button.grid(row=0, column=2, sticky="news", padx=20, pady=10)

    #Frame de los botones
    auxiliar_frame = customtkinter.CTkFrame(frame,fg_color="#D9D9D9")
    auxiliar_frame.grid(row=3, column=0, padx=20, pady=10)


    #Boton para mostrar la grafica de Gantt
    graph_button = customtkinter.CTkButton(auxiliar_frame, text="Graficar ", height=40,font=('Helvetica', 20),fg_color='#284EEE', command=Diagrama_gantt)#284EEE,#Originak=3A7EBF,nueva=#0018BC
    graph_button.grid(row=0, column=0, sticky="news", padx=20, pady=10)
    
    volver_btn = customtkinter.CTkButton(auxiliar_frame, text="PERT-CPM", height=40,font=('Helvetica', 20),fg_color='#3A7EBF', command=lambda: [window.destroy(), PERT_CPM_button_clicked()])
    volver_btn.grid(row=0, column=2, sticky="news", padx=20, pady=10)

    # Botón para salir
    buttonSalir = customtkinter.CTkButton(auxiliar_frame, text="Salir", height=40, font=('Helvetica', 20, 'bold'), fg_color='red', hover_color='#B80507', command=window.destroy)
    buttonSalir.grid(row=1, column=1, padx=20, pady=(10, 10))

    window.mainloop()


def Diagrama_gantt():
    df = pd.read_excel("C:/ #/ #/ #/ #/ #/Interfaz/ExcelGantt.xlsx")

    #Asegurarse de que las fechas estén en el formato correcto
    df['Start'] = pd.to_datetime(df['Start'], format='%d/%m/%Y')  #
    df['End'] = pd.to_datetime(df['End'], format='%d/%m/%Y')      

    #Eliminar filas con valores NaN
    df.dropna(subset=['Task', 'Department', 'Start', 'End', 'Completion'], inplace=True)

    def generate_random_color():
        return '#{:06x}'.format(random.randint(0, 0xFFFFFF))

    #Diccionario para almacenar colores 
    color_dict = {}

    #Asignar colores únicos a cada responsable
    def color(row):
        if row.Department not in color_dict:
            color_dict[row.Department] = generate_random_color()
        return color_dict[row.Department]

    df['color'] = df.apply(color, axis=1)

    proj_start = df.Start.min()

    df['start_num'] = (df.Start - proj_start).dt.days
    df['end_num'] = (df.End - proj_start).dt.days
    df['days_start_to_end'] = df.end_num - df.start_num

    df['current_num'] = df['days_start_to_end'] * df['Completion']

    plt.close('all')
    fig, ax = plt.subplots(1, figsize=(20, 5))

    ax.barh(df.Task.astype(str), df.current_num, left=df.start_num, color=df.color)
    ax.barh(df.Task.astype(str), df.days_start_to_end, left=df.start_num, color=df.color, alpha=0.5)

    for idx, row in df.iterrows():
        #Agregar porcentaje
        if not pd.isna(row.Completion):
            ax.text(row.end_num + 0.1, idx, f"{int(row.Completion * 100)}%", va='center', alpha=0.8)

    #Agregar leyenda de responsables y colores
    unique_departments = df.drop_duplicates(subset=['Department'])
    for idx, row in unique_departments.iterrows():
        ax.scatter([], [], color=row['color'], label=row['Department'])

    #Agregar leyenda al final del gráfico
    ax.legend(loc='upper left', bbox_to_anchor=(1, 1), title='Responsables', markerscale=2)
    
    #Establecer los limites del gráfico
    ax.set_xlim([0, df.end_num.max() + 1])
    ax.set_ylim([-0.5, len(df) - 0.5])
    
    xticks = np.arange(0, df.end_num.max() + 1, 1) #Cambiar el intervalo a 1 día(eje y)
    xticks_labels = pd.date_range(proj_start, end=df.End.max()).strftime("%m/%d")
    ax.set_xticks(xticks)
    ax.set_xticklabels(xticks_labels, rotation=45)  
    plt.gca().xaxis.grid(True)
    plt.title("Diagrama de Gantt")
    plt.show()